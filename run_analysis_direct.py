"""
openpyxl로 직접 Excel 읽어서 분석하는 스크립트
pandas 버전 충돌 우회
"""

import sys
from pathlib import Path
from collections import defaultdict

# openpyxl만 import
try:
    import openpyxl
except ImportError:
    print("✗ openpyxl이 설치되지 않았습니다.")
    print("  pip install --user openpyxl")
    sys.exit(1)

print("=" * 80)
print("요관 결석 탐지 AI 성능 분석 - 실제 데이터 분석")
print("=" * 80)

# Excel 파일 로딩
print("\n[1] Excel 파일 구조 확인...")
print("-" * 80)

def load_excel_structure(file_path):
    """Excel 파일의 구조를 확인합니다"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 컬럼명 (첫 번째 행)
    columns = [cell.value for cell in ws[1]]

    # 데이터 샘플 (2-4행)
    data_sample = []
    for row in ws.iter_rows(min_row=2, max_row=4, values_only=True):
        data_sample.append(row)

    # 총 행 수
    total_rows = ws.max_row - 1  # 헤더 제외

    return {
        'columns': columns,
        'sample': data_sample,
        'total_rows': total_rows,
        'workbook': wb
    }

files = {
    'BCR': 'BCR_result.xlsx',
    'EMS': 'EMS_result.xlsx',
    'Resident': 'Resident_result.xlsx'
}

file_info = {}

for reader_type, filename in files.items():
    try:
        info = load_excel_structure(filename)
        file_info[reader_type] = info

        print(f"\n[{reader_type}] {filename}")
        print(f"  총 행 수: {info['total_rows']:,}개")
        print(f"  컬럼 수: {len(info['columns'])}개")
        print(f"  컬럼명: {info['columns']}")
        print(f"\n  데이터 샘플 (처음 3행):")
        for i, row in enumerate(info['sample'], 2):
            print(f"    Row {i}: {row}")

    except Exception as e:
        print(f"\n✗ {filename} 로딩 실패: {e}")

# 공통 컬럼 구조 분석
if file_info:
    print("\n" + "=" * 80)
    print("[2] 데이터 구조 분석")
    print("=" * 80)

    # 모든 파일의 컬럼이 동일한지 확인
    all_columns = [tuple(info['columns']) for info in file_info.values()]

    if len(set(all_columns)) == 1:
        print("\n✓ 모든 파일이 동일한 컬럼 구조를 가집니다")
        print(f"  공통 컬럼: {list(all_columns[0])}")
    else:
        print("\n⚠ 파일마다 컬럼 구조가 다릅니다")
        for reader_type, info in file_info.items():
            print(f"  [{reader_type}]: {info['columns']}")

    # 총 레코드 수
    total_records = sum(info['total_rows'] for info in file_info.values())
    print(f"\n📊 전체 레코드 수: {total_records:,}개")
    for reader_type, info in file_info.items():
        pct = (info['total_rows'] / total_records * 100) if total_records > 0 else 0
        print(f"  [{reader_type}]: {info['total_rows']:,}개 ({pct:.1f}%)")

# 실제 분석을 위한 준비
print("\n" + "=" * 80)
print("[3] 분석 준비 상태")
print("=" * 80)

if file_info:
    # 첫 번째 파일의 컬럼으로 필요한 컬럼 확인
    sample_columns = file_info[list(file_info.keys())[0]]['columns']

    print("\n현재 데이터 컬럼:")
    for i, col in enumerate(sample_columns, 1):
        print(f"  {i}. {col}")

    print("\n분석에 필요한 컬럼:")
    required = [
        "patient_id (환자 ID)",
        "mode (assisted/unaided)",
        "ground_truth (실제 병변 유무)",
        "prediction (예측 결과)"
    ]
    for req in required:
        print(f"  - {req}")

    print("\n💡 다음 단계:")
    print("  1. 실제 Excel 파일의 컬럼을 위 목록과 매핑")
    print("  2. data_loader.py에서 컬럼 매핑 로직 추가")
    print("  3. 또는 Excel 파일의 컬럼명을 표준화")

print("\n" + "=" * 80)
print("분석 준비 완료!")
print("=" * 80)
