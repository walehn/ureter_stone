#!/usr/bin/env python3
"""
요관 결석 탐지 AI 성능 분석 - 통합 파이프라인

전체 분석 프로세스를 하나의 명령으로 실행합니다.

Usage:
    python main.py [--skip-bootstrap] [--skip-visualization] [--quick]

Options:
    --skip-bootstrap      Bootstrap 분석 건너뛰기 (시간 절약, ~5분)
    --skip-visualization  시각화 건너뛰기
    --quick               빠른 실행 (Bootstrap B=100, 시각화/보고서 스킵)
    --help               도움말 출력
"""

import sys
import argparse
from pathlib import Path
from datetime import datetime
import time

# 프로젝트 루트를 path에 추가
sys.path.insert(0, str(Path(__file__).parent))

import openpyxl
from src.patient_metrics import PatientLevelAnalyzer
from src.bootstrap import BootstrapAnalyzer
from src.gee_analysis import GEEAnalyzer
from src.dca import DecisionCurveAnalyzer
from src.lesion_metrics import LesionMetricsCalculator
from src.visualization import Visualizer
from src.reporter import SupplementReporter
from src.logger import setup_logger

logger = setup_logger("main_pipeline", level="INFO")


class AnalysisPipeline:
    """
    전체 분석 파이프라인 관리 클래스
    """

    def __init__(self,
                 skip_bootstrap: bool = False,
                 skip_visualization: bool = False,
                 quick_mode: bool = False):
        """
        초기화

        Args:
            skip_bootstrap: Bootstrap 분석 건너뛰기
            skip_visualization: 시각화 건너뛰기
            quick_mode: 빠른 실행 모드 (Bootstrap B=100, 시각화 스킵)
        """
        self.skip_bootstrap = skip_bootstrap or quick_mode
        self.skip_visualization = skip_visualization or quick_mode
        self.quick_mode = quick_mode

        self.results_dir = Path("results")
        self.results_dir.mkdir(parents=True, exist_ok=True)

        # 분석 시작 시간
        self.start_time = datetime.now()

        logger.info("=" * 80)
        logger.info("요관 결석 탐지 AI 성능 분석 파이프라인")
        logger.info("=" * 80)
        logger.info(f"시작 시간: {self.start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"실행 모드: {'Quick' if quick_mode else 'Full'}")
        logger.info(f"Bootstrap: {'SKIP' if self.skip_bootstrap else 'RUN'}")
        logger.info(f"Visualization: {'SKIP' if self.skip_visualization else 'RUN'}")
        logger.info("=" * 80)

    def step1_load_data(self):
        """Step 1: 데이터 로딩 및 검증"""
        logger.info("\n[Step 1/8] 데이터 로딩 및 검증...")
        logger.info("-" * 80)

        readers_config = {
            'BCR': {'file': 'BCR_result.xlsx', 'result_idx': (18, 19)},
            'EMS': {'file': 'EMS_result.xlsx', 'result_idx': (18, 19)},
            'Resident': {'file': 'Resident_result.xlsx', 'result_idx': (19, 20)}
        }

        self.data = {}

        for reader, config in readers_config.items():
            logger.info(f"  Loading {reader}: {config['file']}")

            wb = openpyxl.load_workbook(config['file'], data_only=True)
            ws = wb.active

            patient_data = []
            for row_idx in range(2, ws.max_row + 1):
                pid = ws.cell(row_idx, 3 if reader == 'Resident' else 2).value
                result_unaided = ws.cell(row_idx, config['result_idx'][0]).value
                result_assisted = ws.cell(row_idx, config['result_idx'][1]).value

                if pid is not None:
                    patient_data.append({
                        'patient_id': pid,
                        'result_unaided': result_unaided,
                        'result_assisted': result_assisted
                    })

            self.data[reader] = patient_data
            logger.info(f"  ✓ {reader}: {len(patient_data)} patients loaded")

        logger.info("✓ Step 1 완료: 모든 데이터 로딩 성공")
        return True

    def step2_patient_metrics(self):
        """Step 2: Patient-level Metrics 계산"""
        logger.info("\n[Step 2/8] Patient-level Performance Metrics 계산...")
        logger.info("-" * 80)

        analyzer = PatientLevelAnalyzer()
        all_results = {}

        for reader, data in self.data.items():
            logger.info(f"  Analyzing {reader}...")
            results = analyzer.analyze(data)
            all_results[reader] = results

            # 간단한 요약 출력
            assisted = results['assisted']['metrics']
            unaided = results['unaided']['metrics']
            logger.info(f"    Assisted:  Se={assisted['sensitivity']:.3f}, Sp={assisted['specificity']:.3f}")
            logger.info(f"    Unaided:   Se={unaided['sensitivity']:.3f}, Sp={unaided['specificity']:.3f}")

        # 결과 저장
        import json
        output_file = self.results_dir / "analysis_results.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(all_results, f, indent=2, ensure_ascii=False)

        logger.info(f"✓ Step 2 완료: {output_file}")
        return True

    def step3_bootstrap(self):
        """Step 3: Bootstrap Analysis"""
        if self.skip_bootstrap:
            logger.info("\n[Step 3/8] Bootstrap Analysis... SKIPPED")
            return True

        logger.info("\n[Step 3/8] Bootstrap Analysis (Patient-level resampling)...")
        logger.info("-" * 80)

        n_iterations = 100 if self.quick_mode else 1000
        analyzer = BootstrapAnalyzer(n_iterations=n_iterations, random_seed=42)

        for reader, data in self.data.items():
            logger.info(f"  {reader}: Running B={n_iterations} iterations...")
            start = time.time()

            results = analyzer.analyze(data)

            elapsed = time.time() - start
            logger.info(f"  ✓ {reader}: {elapsed:.1f}초 소요")

            # 결과 저장
            output_dir = self.results_dir / "bootstrap" / reader
            output_dir.mkdir(parents=True, exist_ok=True)

            analyzer.save_results(results, output_dir)
            logger.info(f"    → {output_dir}/")

        logger.info("✓ Step 3 완료: Bootstrap Analysis")
        return True

    def step4_gee(self):
        """Step 4: GEE Analysis"""
        logger.info("\n[Step 4/8] GEE Analysis (Cluster-robust inference)...")
        logger.info("-" * 80)

        analyzer = GEEAnalyzer()

        for reader, data in self.data.items():
            logger.info(f"  {reader}: Fitting GEE model...")

            results = analyzer.analyze(data)

            # OR 출력
            coef = results['coefficients']['Mode (Assisted vs Unaided)']
            or_val = coef['OR']
            p_val = coef['p_value']
            sig = "***" if p_val < 0.001 else "**" if p_val < 0.01 else "*" if p_val < 0.05 else "ns"

            logger.info(f"    OR = {or_val:.3f}, p = {p_val:.4f} {sig}")

            # 결과 저장
            output_dir = self.results_dir / "gee" / reader
            output_dir.mkdir(parents=True, exist_ok=True)

            analyzer.save_results(results, output_dir)
            logger.info(f"    → {output_dir}/")

        logger.info("✓ Step 4 완료: GEE Analysis")
        return True

    def step5_dca(self):
        """Step 5: Decision Curve Analysis"""
        logger.info("\n[Step 5/8] Decision Curve Analysis...")
        logger.info("-" * 80)

        analyzer = DecisionCurveAnalyzer(threshold_min=0.05, threshold_max=0.25, n_thresholds=50)

        for reader, data in self.data.items():
            logger.info(f"  {reader}: Computing net benefit across thresholds...")

            results = analyzer.analyze(data)

            # 최대 delta 출력
            max_delta = max(results['delta_net_benefit'])
            max_idx = results['delta_net_benefit'].index(max_delta)
            max_threshold = results['thresholds'][max_idx]

            logger.info(f"    Max Δ NB = {max_delta:+.4f} at threshold = {max_threshold:.3f}")

            # 결과 저장
            output_dir = self.results_dir / "dca" / reader
            output_dir.mkdir(parents=True, exist_ok=True)

            analyzer.save_results(results, output_dir)
            logger.info(f"    → {output_dir}/")

        logger.info("✓ Step 5 완료: Decision Curve Analysis")
        return True

    def step6_lesion_metrics(self):
        """Step 6: Lesion-level Detection Metrics"""
        logger.info("\n[Step 6/8] Lesion-level Detection Metrics...")
        logger.info("-" * 80)

        import openpyxl

        readers_config = {
            'BCR': {'file': 'BCR_result.xlsx', 'tp_idx': (4, 7), 'fp_idx': (5, 8), 'fn_idx': (6, 9)},
            'EMS': {'file': 'EMS_result.xlsx', 'tp_idx': (4, 7), 'fp_idx': (5, 8), 'fn_idx': (6, 9)},
            'Resident': {'file': 'Resident_result.xlsx', 'tp_idx': (5, 8), 'fp_idx': (6, 9), 'fn_idx': (7, 10)}
        }

        calculator = LesionMetricsCalculator()

        for reader, config in readers_config.items():
            logger.info(f"  {reader}: Aggregating lesion counts...")

            wb = openpyxl.load_workbook(config['file'], data_only=True)
            ws = wb.active

            # 병변 수 집계
            tp_unaided = sum(ws.cell(row, config['tp_idx'][0]).value or 0 for row in range(2, ws.max_row + 1))
            fp_unaided = sum(ws.cell(row, config['fp_idx'][0]).value or 0 for row in range(2, ws.max_row + 1))
            fn_unaided = sum(ws.cell(row, config['fn_idx'][0]).value or 0 for row in range(2, ws.max_row + 1))

            tp_assisted = sum(ws.cell(row, config['tp_idx'][1]).value or 0 for row in range(2, ws.max_row + 1))
            fp_assisted = sum(ws.cell(row, config['fp_idx'][1]).value or 0 for row in range(2, ws.max_row + 1))
            fn_assisted = sum(ws.cell(row, config['fn_idx'][1]).value or 0 for row in range(2, ws.max_row + 1))

            # 메트릭 계산
            results = calculator.calculate_comparison(
                tp_assisted, fp_assisted, fn_assisted,
                tp_unaided, fp_unaided, fn_unaided
            )

            # 요약 출력
            delta = results['delta']
            logger.info(f"    Δ Precision = {delta['precision']:+.1%}, Δ Recall = {delta['recall']:+.1%}")

            # 결과 저장
            output_dir = self.results_dir / "lesion_metrics" / reader
            output_dir.mkdir(parents=True, exist_ok=True)

            calculator.save_results(results, output_dir)
            logger.info(f"    → {output_dir}/")

        logger.info("✓ Step 6 완료: Lesion-level Metrics")
        return True

    def step7_visualization(self):
        """Step 7: Visualization"""
        if self.skip_visualization:
            logger.info("\n[Step 7/8] Visualization... SKIPPED")
            return True

        logger.info("\n[Step 7/8] Visualization (300 dpi PNG)...")
        logger.info("-" * 80)

        # run_visualization.py를 직접 실행하는 대신 간소화된 버전
        logger.info("  Generating graphs for all readers...")

        viz = Visualizer(dpi=300)
        readers = ['BCR', 'EMS', 'Resident']
        output_base = self.results_dir / "figures"
        output_base.mkdir(parents=True, exist_ok=True)

        graph_count = 0

        # 각 리더별 주요 그래프만 생성
        import json
        import csv

        for reader in readers:
            reader_output = output_base / reader
            reader_output.mkdir(parents=True, exist_ok=True)

            # 1. DCA
            try:
                dca_file = self.results_dir / f"dca/{reader}/dca_curve.csv"
                if dca_file.exists():
                    import numpy as np
                    thresholds, nb_a, nb_u, nb_all, nb_none = [], [], [], [], []
                    with open(dca_file, 'r') as f:
                        for row in csv.DictReader(f):
                            thresholds.append(float(row['Threshold']))
                            nb_a.append(float(row['NB_Assisted']))
                            nb_u.append(float(row['NB_Unaided']))
                            nb_all.append(float(row['NB_Treat_All']))
                            nb_none.append(float(row['NB_Treat_None']))

                    viz.plot_decision_curve(
                        np.array(thresholds), np.array(nb_a), np.array(nb_u),
                        np.array(nb_all), np.array(nb_none),
                        reader_output / "decision_curve.png",
                        f"{reader} - Decision Curve Analysis"
                    )
                    graph_count += 1
            except Exception as e:
                logger.warning(f"    DCA visualization failed for {reader}: {e}")

            # 2. Patient-level metrics
            try:
                patient_file = self.results_dir / "analysis_results.json"
                with open(patient_file, 'r') as f:
                    patient_data = json.load(f)

                reader_data = patient_data[reader]
                metrics_a = reader_data['assisted']['metrics']
                metrics_u = reader_data['unaided']['metrics']

                viz.plot_metrics_comparison(
                    metrics_a, metrics_u,
                    ['sensitivity', 'specificity', 'ppv', 'npv'],
                    reader_output / "patient_metrics.png",
                    f"{reader} - Patient-level Metrics"
                )
                graph_count += 1
            except Exception as e:
                logger.warning(f"    Patient metrics visualization failed for {reader}: {e}")

            # 3. Lesion-level metrics
            try:
                lesion_file = self.results_dir / f"lesion_metrics/{reader}/lesion_metrics.json"
                with open(lesion_file, 'r') as f:
                    lesion_data = json.load(f)

                viz.plot_metrics_comparison(
                    lesion_data['assisted'], lesion_data['unaided'],
                    ['precision', 'recall', 'f1_score'],
                    reader_output / "lesion_metrics.png",
                    f"{reader} - Lesion-level Metrics"
                )
                graph_count += 1
            except Exception as e:
                logger.warning(f"    Lesion metrics visualization failed for {reader}: {e}")

        logger.info(f"  ✓ Generated {graph_count} graphs")
        logger.info(f"✓ Step 7 완료: Visualization → {output_base}/")
        return True

    def step8_reporting(self):
        """Step 8: Generate Final Report"""
        logger.info("\n[Step 8/8] Generating Supplement-ready Report...")
        logger.info("-" * 80)

        reporter = SupplementReporter(results_dir=self.results_dir)
        report_path = reporter.generate_full_report()

        logger.info(f"✓ Step 8 완료: Report → {report_path}")
        return True

    def run(self):
        """전체 파이프라인 실행"""
        try:
            # 각 단계 실행
            self.step1_load_data()
            self.step2_patient_metrics()
            self.step3_bootstrap()
            self.step4_gee()
            self.step5_dca()
            self.step6_lesion_metrics()
            self.step7_visualization()
            self.step8_reporting()

            # 완료 메시지
            end_time = datetime.now()
            elapsed = (end_time - self.start_time).total_seconds()

            logger.info("\n" + "=" * 80)
            logger.info("파이프라인 완료!")
            logger.info("=" * 80)
            logger.info(f"총 소요 시간: {elapsed:.1f}초 ({elapsed/60:.1f}분)")
            logger.info(f"종료 시간: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
            logger.info("\n결과 위치:")
            logger.info(f"  - 전체 결과: {self.results_dir}/")
            logger.info(f"  - 최종 보고서: {self.results_dir}/reports/supplement_full_report.md")
            logger.info(f"  - 그래프: {self.results_dir}/figures/")
            logger.info("=" * 80)

            return True

        except Exception as e:
            logger.error(f"\n파이프라인 실행 중 오류 발생: {e}")
            import traceback
            traceback.print_exc()
            return False


def main():
    """메인 함수"""
    parser = argparse.ArgumentParser(
        description='요관 결석 탐지 AI 성능 분석 - 통합 파이프라인',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예제:
  python main.py                      # 전체 분석 실행 (Bootstrap B=1000 포함)
  python main.py --skip-bootstrap     # Bootstrap 건너뛰기 (~5분 절약)
  python main.py --quick              # 빠른 실행 (Bootstrap B=100, 시각화 스킵)
        """
    )

    parser.add_argument('--skip-bootstrap', action='store_true',
                        help='Bootstrap 분석 건너뛰기 (시간 절약)')
    parser.add_argument('--skip-visualization', action='store_true',
                        help='시각화 건너뛰기')
    parser.add_argument('--quick', action='store_true',
                        help='빠른 실행 모드 (Bootstrap B=100, 시각화 스킵)')

    args = parser.parse_args()

    # 파이프라인 실행
    pipeline = AnalysisPipeline(
        skip_bootstrap=args.skip_bootstrap,
        skip_visualization=args.skip_visualization,
        quick_mode=args.quick
    )

    success = pipeline.run()
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
