"""
실제 Excel 데이터를 읽어서 FR-01, FR-02 분석 실행
"""

import openpyxl
from collections import defaultdict
import json
import csv
from pathlib import Path

print("=" * 80)
print("요관 결석 탐지 AI 성능 분석 - 3개 리더 비교 분석")
print("=" * 80)

def load_reader_data(filename):
    """Excel 파일에서 실제 데이터 로딩 (Row 4=헤더, Row 5~=데이터)"""
    # data_only=True로 Excel 수식의 계산된 값을 가져옴
    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.active

    # Row 4가 실제 컬럼명
    header_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]

    # 데이터 추출 (Row 5부터)
    data = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        data.append(row)

    return header_row, data


def analyze_patient_level(data, header, reader_type):
    """
    환자 단위 성능 분석

    컬럼 구조:
    - BCR/EMS: PID=Col 2, Result(Unaided)=Col 18, Result(Assisted)=Col 19
    - Resident: PID=Col 3, Result(Unaided)=Col 19, Result(Assisted)=Col 20
    """

    # Reader 타입에 따라 인덱스 설정
    if reader_type == 'Resident':
        pid_idx = 3
        result_unaided_idx = 19
        result_assisted_idx = 20
    else:  # BCR, EMS
        pid_idx = 2
        result_unaided_idx = 18
        result_assisted_idx = 19

    patient_data = defaultdict(lambda: {
        'unaided': {'result': None},
        'assisted': {'result': None}
    })

    for row in data:
        if not row or len(row) <= result_assisted_idx:
            continue

        if row[pid_idx] is None:
            continue

        patient_id = str(row[pid_idx])

        # Result 컬럼 읽기
        result_unaided = row[result_unaided_idx]
        result_assisted = row[result_assisted_idx]

        # 환자 ID당 하나의 결과만 저장 (첫 번째 행 우선)
        if patient_data[patient_id]['unaided']['result'] is None:
            patient_data[patient_id]['unaided']['result'] = result_unaided
            patient_data[patient_id]['assisted']['result'] = result_assisted

    return patient_data


# 3개 리더 데이터 로딩
readers = {
    'BCR': 'BCR_result.xlsx',
    'EMS': 'EMS_result.xlsx',
    'Resident': 'Resident_result.xlsx'
}

results = {}

print("\n[1] 데이터 로딩...")
print("-" * 80)

for reader_name, filename in readers.items():
    try:
        header, data = load_reader_data(filename)
        print(f"\n[{reader_name}]")
        print(f"  파일: {filename}")
        print(f"  환자 수: {len(data):,}명")
        print(f"  컬럼: {len(header)}개")

        # 실제 컬럼명 출력 (None 제외)
        real_columns = [col for col in header if col is not None]
        print(f"  주요 컬럼: {', '.join(real_columns[:10])}")

        # Patient-level 분석
        patient_data = analyze_patient_level(data, header, reader_name)

        if patient_data:
            print(f"  분석된 환자: {len(patient_data)}명")

            # 첫 5명 환자 샘플 출력
            print(f"\n  환자 데이터 샘플 (첫 5명):")
            for i, (pid, pdata) in enumerate(list(patient_data.items())[:5], 1):
                unaided_result = pdata['unaided']['result']
                assisted_result = pdata['assisted']['result']
                print(f"    {i}. PID={pid}: Unaided={unaided_result}, Assisted={assisted_result}")

        results[reader_name] = patient_data

    except Exception as e:
        print(f"\n✗ {filename} 처리 실패: {e}")
        import traceback
        traceback.print_exc()


# 성능 지표 계산
print("\n" + "=" * 80)
print("[2] 환자단위 성능 지표 계산")
print("=" * 80)

def calculate_confusion_matrix(patient_data, mode='assisted'):
    """
    Confusion Matrix 계산
    Result 컬럼 값: 'TP', 'FP', 'TN', 'FN'
    """
    cm = {'TP': 0, 'FP': 0, 'TN': 0, 'FN': 0}

    for pid, pdata in patient_data.items():
        result = pdata[mode]['result']
        if result in cm:
            cm[result] += 1

    return cm


def calculate_metrics(cm):
    """성능 지표 계산"""
    tp, fp, tn, fn = cm['TP'], cm['FP'], cm['TN'], cm['FN']

    sensitivity = tp / (tp + fn) if (tp + fn) > 0 else 0
    specificity = tn / (tn + fp) if (tn + fp) > 0 else 0
    ppv = tp / (tp + fp) if (tp + fp) > 0 else 0
    npv = tn / (tn + fn) if (tn + fn) > 0 else 0

    return {
        'sensitivity': sensitivity,
        'specificity': specificity,
        'ppv': ppv,
        'npv': npv
    }


# 모든 결과를 저장할 딕셔너리
analysis_results = {}

for reader_name, patient_data in results.items():
    if not patient_data:
        continue

    print(f"\n[{reader_name}]")
    print("-" * 80)

    for mode in ['unaided', 'assisted']:
        mode_name = "Without AI" if mode == 'unaided' else "With AI"

        cm = calculate_confusion_matrix(patient_data, mode)
        metrics = calculate_metrics(cm)

        print(f"\n  {mode_name}:")
        print(f"    환자 수: {len(patient_data)}명")
        print(f"\n    Confusion Matrix:")
        print(f"      TP: {cm['TP']:3d}  FP: {cm['FP']:3d}")
        print(f"      FN: {cm['FN']:3d}  TN: {cm['TN']:3d}")
        print(f"\n    성능 지표:")
        print(f"      Sensitivity: {metrics['sensitivity']:.3f} ({metrics['sensitivity']*100:.1f}%)")
        print(f"      Specificity: {metrics['specificity']:.3f} ({metrics['specificity']*100:.1f}%)")
        print(f"      PPV:         {metrics['ppv']:.3f} ({metrics['ppv']*100:.1f}%)")
        print(f"      NPV:         {metrics['npv']:.3f} ({metrics['npv']*100:.1f}%)")

    # Delta 계산
    cm_unaided = calculate_confusion_matrix(patient_data, 'unaided')
    cm_assisted = calculate_confusion_matrix(patient_data, 'assisted')

    metrics_unaided = calculate_metrics(cm_unaided)
    metrics_assisted = calculate_metrics(cm_assisted)

    deltas = {}
    print(f"\n  Δ (With AI - Without AI):")
    for metric in ['sensitivity', 'specificity', 'ppv', 'npv']:
        delta = metrics_assisted[metric] - metrics_unaided[metric]
        deltas[f'delta_{metric}'] = delta
        direction = "↑" if delta > 0 else "↓" if delta < 0 else "="
        print(f"    {metric.upper():12s}: {delta:+.3f} ({delta*100:+.1f}%) {direction}")

    # 결과 저장
    analysis_results[reader_name] = {
        'n_patients': len(patient_data),
        'unaided': {
            'confusion_matrix': cm_unaided,
            'metrics': metrics_unaided
        },
        'assisted': {
            'confusion_matrix': cm_assisted,
            'metrics': metrics_assisted
        },
        'deltas': deltas
    }


# 결과 저장
print("\n" + "=" * 80)
print("[3] 결과 저장")
print("=" * 80)

results_dir = Path("results")
results_dir.mkdir(exist_ok=True)

# 1. JSON 저장 (전체 구조화 데이터)
json_file = results_dir / "analysis_results.json"
with open(json_file, 'w', encoding='utf-8') as f:
    json.dump(analysis_results, f, indent=2, ensure_ascii=False)
print(f"\n✓ JSON 저장: {json_file}")

# 2. CSV 저장 (비교 테이블)
csv_file = results_dir / "comparison_table.csv"
with open(csv_file, 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)

    # 헤더
    writer.writerow(['Reader', 'Mode', 'N_Patients', 'TP', 'FP', 'FN', 'TN',
                     'Sensitivity', 'Specificity', 'PPV', 'NPV'])

    # 데이터
    for reader_name, data in analysis_results.items():
        for mode in ['unaided', 'assisted']:
            mode_name = 'Without AI' if mode == 'unaided' else 'With AI'
            cm = data[mode]['confusion_matrix']
            m = data[mode]['metrics']

            writer.writerow([
                reader_name,
                mode_name,
                data['n_patients'],
                cm['TP'], cm['FP'], cm['FN'], cm['TN'],
                f"{m['sensitivity']:.3f}",
                f"{m['specificity']:.3f}",
                f"{m['ppv']:.3f}",
                f"{m['npv']:.3f}"
            ])

print(f"✓ CSV 저장: {csv_file}")

# 3. Markdown 리포트 저장
md_file = results_dir / "analysis_report.md"
with open(md_file, 'w', encoding='utf-8') as f:
    f.write("# 요관 결석 탐지 AI 성능 분석 결과\n\n")
    f.write("## 3개 리더 비교 분석\n\n")

    for reader_name, data in analysis_results.items():
        f.write(f"### {reader_name}\n\n")
        f.write(f"**환자 수**: {data['n_patients']}명\n\n")

        # Without AI
        f.write("#### Without AI\n\n")
        cm_u = data['unaided']['confusion_matrix']
        m_u = data['unaided']['metrics']
        f.write(f"- **Confusion Matrix**: TP={cm_u['TP']}, FP={cm_u['FP']}, FN={cm_u['FN']}, TN={cm_u['TN']}\n")
        f.write(f"- **Sensitivity**: {m_u['sensitivity']:.3f} ({m_u['sensitivity']*100:.1f}%)\n")
        f.write(f"- **Specificity**: {m_u['specificity']:.3f} ({m_u['specificity']*100:.1f}%)\n")
        f.write(f"- **PPV**: {m_u['ppv']:.3f} ({m_u['ppv']*100:.1f}%)\n")
        f.write(f"- **NPV**: {m_u['npv']:.3f} ({m_u['npv']*100:.1f}%)\n\n")

        # With AI
        f.write("#### With AI\n\n")
        cm_a = data['assisted']['confusion_matrix']
        m_a = data['assisted']['metrics']
        f.write(f"- **Confusion Matrix**: TP={cm_a['TP']}, FP={cm_a['FP']}, FN={cm_a['FN']}, TN={cm_a['TN']}\n")
        f.write(f"- **Sensitivity**: {m_a['sensitivity']:.3f} ({m_a['sensitivity']*100:.1f}%)\n")
        f.write(f"- **Specificity**: {m_a['specificity']:.3f} ({m_a['specificity']*100:.1f}%)\n")
        f.write(f"- **PPV**: {m_a['ppv']:.3f} ({m_a['ppv']*100:.1f}%)\n")
        f.write(f"- **NPV**: {m_a['npv']:.3f} ({m_a['npv']*100:.1f}%)\n\n")

        # Delta
        f.write("#### Δ (With AI - Without AI)\n\n")
        for metric in ['sensitivity', 'specificity', 'ppv', 'npv']:
            delta = data['deltas'][f'delta_{metric}']
            direction = "↑" if delta > 0 else "↓" if delta < 0 else "="
            f.write(f"- **{metric.upper()}**: {delta:+.3f} ({delta*100:+.1f}%) {direction}\n")
        f.write("\n---\n\n")

print(f"✓ Markdown 저장: {md_file}")

print("\n" + "=" * 80)
print("분석 완료!")
print("=" * 80)
